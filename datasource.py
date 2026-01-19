# preenchimento_app/datasource.py
import os
import csv
from openpyxl import load_workbook


def load_table(file_path: str, sheet_name: str | None = None):
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            rows = [dict(r) for r in reader]
        headers = list(rows[0].keys()) if rows else []
        return headers, rows

    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        wb = load_workbook(file_path, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        data = list(ws.iter_rows(values_only=True))
        if not data:
            return [], []

        headers = [str(h).strip() if h is not None else "" for h in data[0]]
        rows = []
        for r in data[1:]:
            row_dict = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                v = r[i] if i < len(r) else ""
                row_dict[h] = "" if v is None else str(v)
            rows.append(row_dict)

        return headers, rows

    raise ValueError("Formato não suportado. Use .csv ou .xlsx")
